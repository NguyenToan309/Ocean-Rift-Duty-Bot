"""
utils/export_utils.py — Logic tạo file CSV/Excel và ký HMAC
Module dùng chung giữa bot/cogs/export.py và web/routers/export.py
"""
import io
import hmac
import hashlib

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from bot.config import settings
from bot.utils.time_utils import minutes_to_hhmm


# Các ký tự "leading" mà Excel/LibreOffice/Numbers diễn dịch là công thức.
# Username Discord do user kiểm soát → attacker có thể đặt tên kiểu
# "=cmd|'/c calc'!A1" để khi staff mở CSV thì shell chạy. Mọi cell text
# user-controlled phải đi qua _sanitize_cell trước khi vào DataFrame.
_CSV_INJECTION_PREFIX = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value):
    """Prepend dấu ' nếu chuỗi bắt đầu bằng ký tự công thức.

    Áp dụng cho mọi cell text có thể chứa input user (username, period_label,
    raw_text…). Giá trị non-string trả nguyên (số, datetime đã format từ trước).
    """
    if not isinstance(value, str):
        return value
    if value.startswith(_CSV_INJECTION_PREFIX):
        return "'" + value
    return value

# Mapping tên cột kỹ thuật → tên hiển thị tiếng Việt
EXPORT_COLUMNS: dict[str, str] = {
    "guild_id":        "Guild ID",
    "guild_name":      "Tên Server",
    "user_id":         "User ID",
    "username":        "Tên thành viên",
    "started_at":      "Giờ bắt đầu",
    "ended_at":        "Giờ kết thúc",
    "duration_minutes":"Thời gian (phút)",
    "duration_hhmm":   "Thời gian (giờ:phút)",
    "day":             "Ngày",
    "week_number":     "Tuần (ISO)",
    "month":           "Tháng",
    "quarter":         "Quý",
    "year":            "Năm",
}


def logs_to_dataframe(logs: list, guild_name: str) -> pd.DataFrame:
    """
    Chuyển list DutyLog ORM objects → DataFrame chuẩn.
    logs: list của DutyLog instances.
    """
    safe_guild_name = _sanitize_cell(guild_name)
    rows = []
    for log in logs:
        dt = log.started_at
        rows.append({
            "guild_id":         log.guild_id,
            "guild_name":       safe_guild_name,
            "user_id":          log.user_id,
            "username":         _sanitize_cell(log.username),
            "started_at":       dt.strftime("%d/%m/%Y %H:%M:%S"),
            "ended_at":         log.ended_at.strftime("%d/%m/%Y %H:%M:%S"),
            "duration_minutes": log.duration_minutes,
            "duration_hhmm":    minutes_to_hhmm(log.duration_minutes),
            "day":              dt.strftime("%d/%m/%Y"),
            "week_number":      dt.isocalendar()[1],
            "month":            dt.month,
            "quarter":          (dt.month - 1) // 3 + 1,
            "year":             dt.year,
        })

    df = pd.DataFrame(rows, columns=list(EXPORT_COLUMNS.keys()))
    # Đổi sang tên cột tiếng Việt để hiển thị
    df.columns = list(EXPORT_COLUMNS.values())
    return df


def generate_csv_bytes(df: pd.DataFrame) -> bytes:
    """UTF-8 với BOM để Excel Windows mở không bị lỗi tiếng Việt"""
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()


def generate_excel_bytes(df: pd.DataFrame, period_label: str) -> bytes:
    """Excel với header màu Discord blurple, stripe rows, auto-fit column width"""
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Duty Log", startrow=1)
        ws = writer.sheets["Duty Log"]

        # Dòng tiêu đề — period_label có thể chứa input user (date_from/date_to),
        # sanitize để chặn CSV/Excel formula injection.
        safe_label = _sanitize_cell(period_label)
        title_cell = ws.cell(row=1, column=1, value=f"Báo cáo LOG DUTY — {safe_label}")
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="5865F2")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws.row_dimensions[1].height = 28

        # Header row (row 2)
        header_fill = PatternFill("solid", fgColor="4F545C")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Stripe rows xen kẽ màu
        light_stripe = PatternFill("solid", fgColor="EEF0FF")
        for row_idx in range(3, ws.max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = light_stripe

        # Auto-fit độ rộng cột
        for col_idx, col_name in enumerate(df.columns, 1):
            col_values = df.iloc[:, col_idx - 1].astype(str)
            max_len = max(len(str(col_name)), col_values.str.len().max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(int(max_len) + 4, 40)

        # Đóng băng header + tiêu đề
        ws.freeze_panes = "A3"

    return buf.getvalue()


def sign_file(file_bytes: bytes) -> str:
    """HMAC-SHA256 ký nội dung file — dùng để kiểm tra tính toàn vẹn"""
    return hmac.new(
        settings.HMAC_SECRET.encode(),
        file_bytes,
        hashlib.sha256,
    ).hexdigest()
